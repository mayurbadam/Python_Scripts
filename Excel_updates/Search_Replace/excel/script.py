#*********************************************************************************************************************************
# * Name                 : script.py
# * Creation Date        : 25-05-2022
# * Last Modified        : 06-06-2022
#----------------------------------------------------------------------------------------------------------------------------------
# * Author               : Badam Mayur Krishna
# * Author's Email       : mayurkrishna.b@alpha-numero.tech
#----------------------------------------------------------------------------------------------------------------------------------
#* Description           : search and replace strings having one word, multiline, having special characters in xlsx file. 
#**********************************************************************************************************************************'''

#importing libraries
import os
import sys
import openpyxl
import fileinput


#Taking inputs from user for no.of lines, string to be searched and replaced
print ("No.of lines:")
#lines = ''''''
lines = int(input( "> " ))

print ("Text to search for:")
textToSearch = ''''''
for line in range(lines):
   textToSearch += input( "> " )+"\n"

#deleting the last line(\n)
textToSearch = textToSearch[:textToSearch.rfind('\n')]

print ("Text to replace it with:")
textToReplace = ''''''
for line in range(lines):
   textToReplace += input( "> " ) + "\n"
textToReplace = textToReplace[:textToReplace.rfind('\n')]

#loading the file
#print ("File to perform Search-Replace on:")
#fileToSearch  = input( "> " )
#print(fileToSearch)  

wb = openpyxl.load_workbook(r'file.xlsx')
ws = wb.active

#search and replace
for i in range(1, ws.max_row + 1):
  for j in range(1, ws.max_column + 1):
      if ws.cell(i,j).value == textToSearch:
         ws.cell(i,j).value = textToReplace
         print("Found and replaced")
         print(ws.cell(i,j).value)
      else:
         print("Not Found")
         print(ws.cell(i,j).value)

#save workbook
wb.save("file.xlsx")

input( '\n\n Press Enter to exit...' )
