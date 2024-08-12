'''*********************************************************************************************************************************
 * Name                 : script.py
 * Creation Date        : 25-05-2022
 * Last Modified        : 06-06-2022
----------------------------------------------------------------------------------------------------------------------------------
 * Author               : Badam Mayur Krishna
 * Author's Email       : mayurkrishna.b@alpha-numero.tech
----------------------------------------------------------------------------------------------------------------------------------
* Description           : Working with excel
 **********************************************************************************************************************************'''
#import openpyxl library. It offers best features for working with excels using python 
import openpyxl

#Loading excel
book = openpyxl.load_workbook('sample.xlsx')

#Loading sheet which is active
sheet = book.active

#Getting values from the excel and adding into lists
for col1 in sheet.iter_cols(min_row=2, min_col=1, max_row=5, max_col=1):
   alist = []
   for cell1 in col1:
      alist.append(cell1.value)
print(alist)

for col1 in sheet.iter_cols(min_row=2, min_col=2, max_row=5, max_col=2):
   blist = []
   for cell1 in col1:
      blist.append(cell1.value)
print(blist)
print(type(blist[0]))

#initiating the lists
clist = [0,0,0,0] 
dlist = [0,0,0,0] 

#Performing operations
for a in range(0,4):
   clist[a] = int(input("Enter number_" + str(a) + ": "))
   if(alist[a] =='+' or alist[a] == 'Add'):
      dlist[a] = blist[a] + clist[a]
   elif(alist[a] =='-' or alist[a] == 'Subtraction'):
      dlist[a] = blist[a] - clist[a]
   elif(alist[a] =='*' or alist[a] == 'Multiply'):
      dlist[a] = blist[a] * clist[a]
   elif(alist[a] =='/' or alist[a] == 'division'):
      dlist[a] = blist[a] / clist[a]
   sheet.cell(row=(a+2), column=3).value = clist[a]   
   sheet.cell(row=(a+2), column=4).value = dlist[a]   
print(dlist)

#saving the changes made to the excel file
book.save('sample.xlsx')

#OUTPUT:
# ['+', 'Subtraction', '*', 'division']
# [10, 20, 40, 30]
# <class 'int'>
# Enter number_0: 50
# Enter number_1: 12
# Enter number_2: 45 
# Enter number_3: 8
# [60, 8, 1800, 3.75]






#for reference - old code, Not in use
#for x in range(2,5):
#   for col1 in sheet.iter_cols(min_row=2, min_col=1, max_row=5, max_col=1):
#      for cell1 in col1:
#         a = cell1.value
#         x=0
#         for col2 in sheet.iter_cols(min_row=2, min_col=2, max_row=5, max_col=2):
#            clist =[]
#            dlist =[]
#            for cell2 in col2:
#               b = cell2.value
#               b = int(b)
#               c = int(input("Enter number: "))
#
#               if(a =='+' or a == 'Add'):
#                  d = b + c
#               if(a =='-' or a == 'Subtraction'):
#                  d = b - c
#               if(a =='*' or a == 'Multiply'):
#                  d = b * c
#               if(a =='/' or a == 'division'):
#                  d = b / c
#               print(c,d)
#               #sheet.cell(row=int(cell2), column=3).value = c
#               #sheet.cell(row=int(cell2), column=4).value = d
#               #sheet['D2'] = d
#
#               clist.append(c)
#               dlist.append(d)
#               x++;

#x = sheet['C2']
#y = sheet['D2']
#print(x.value)
#print(y.value)
