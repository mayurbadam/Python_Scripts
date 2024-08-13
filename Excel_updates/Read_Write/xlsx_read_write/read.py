'''*********************************************************************************************************************************
 * Name                 : read.py
 * Creation Date        : 25-05-2022
 * Last Modified        : 06-06-2022
----------------------------------------------------------------------------------------------------------------------------------
 * Author               : Badam Mayur Krishna
 * Author's Email       : mayurkrishna.b@alpha-numero.tech
----------------------------------------------------------------------------------------------------------------------------------
* Description           : Reading to following file types - Xlsx file
 **********************************************************************************************************************************'''
#Reading from excel

#importing library openpyxl
import openpyxl

#loading excel using load_wrokbook method
book = openpyxl.load_workbook("b.xlsx")
#print(type(book))

#sheets in the workbook
#sheet = book.sheetnames
#print(sheet)   # print how many sheet available in xls file
#print(book.active.title)   # Current active sheet
#options_1
#s1 = book['Sheet1']

#setting s1 as the sheet which is active
s1 = book.active

#Reading values using different options
print(s1['C3'].value)

print(s1['B2'].value)

print(s1.cell(1,2).value) 

print(s1.cell(row = 3,column = 2).value + "\n")

row = int(s1.max_row)
column = int(s1.max_column)

#Reading all values in the excel using for loop
for i in range(1, row+1):
    for j in range(1, column+1):
        print(s1.cell(i,j).value)








#import openpyxl
#book = openpyxl.load_workbook('b.xlsx')
#
#sheet = book.active
#
##a1 = sheet['C9']
##a2 = sheet['B9']
##a3 = sheet.cell(row=9, column=2)
#a4 = sheet['A1':'C6']
#
##print(a1.value)
##print(a2.value) 
##print(a3.value)
##
##print(type(book))
##print(book)
#
#for c1,c2,c3 in a4:
#    print("{0:8} {1:8} {3:8}".format(c1.value, c2.value, c3.value))
