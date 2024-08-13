'''*********************************************************************************************************************************
 * Name                 : write.py
 * Creation Date        : 27-05-2022
 * Last Modified        : 06-06-2022
----------------------------------------------------------------------------------------------------------------------------------
 * Author               : Badam Mayur Krishna
 * Author's Email       : mayurkrishna.b@alpha-numero.tech
----------------------------------------------------------------------------------------------------------------------------------
* Description           : Writing to following file types - Xlsx file
 **********************************************************************************************************************************'''

#importing Workbook from library openpyxl
from openpyxl import Workbook

#Using Workbook
book = Workbook()
book ['Sheet'].title = "Trainees Details"
#print(book.active)

#assigning s1 with sheet which is active
s1 = book.active

#writing to excel in different ways
s1['A1'].value = "NAME"
s1['B1'].value = "Emp. NO"
s1['C1'].value = "Location"
s1['A2'].value = "Mayur"
s1['B2'].value = "AN215"
s1['C2'].value = "Hyderabad"
s1['A3'].value = "Naga"
s1['B3'].value = "AN214"
s1['C3'].value = "Hyderabad"
s1['A4'].value = "sameeksha"
s1['B4'].value = "AN213"
s1['C4'].value = "Hyderabad"

#saving workbook 
book.save("b.xlsx")

#OUTPUT:
# Hyderabad
# AN215
# Emp. NO
# AN214
# 
# NAME
# Emp. NO
# Location
# Mayur
# AN215
# Hyderabad
# Naga
# AN214
# Hyderabad
# sameeksha
# AN213
# Hyderabad




#from openpyxl import Workbook
#
#book = Workbook()
#sheet = book.active
#
##Writing to excel
##sheet['A1'] = 56
##sheet['A8'] = 43
##
##sheet.cell(row=1, column=2).value = 5
#
#rows = [
#    ["a", 46, 57],
#    ['b', 38, 12],
#    ['c', 59, 78],
#    ['d', 21, 98],
#    ['e', 18, 43],
#    ['f', 15, 67]
#]
#
#for row in rows:
#    sheet.append(row)
#
#book.save("b.xlsx")
#
