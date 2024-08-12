'''*********************************************************************************************************************************
 * Name                 : script.py
 * Creation Date        : 25-05-2022
 * Last Modified        : 06-06-2022
----------------------------------------------------------------------------------------------------------------------------------
 * Author               : Badam Mayur Krishna
 * Author's Email       : mayurkrishna.b@alpha-numero.tech
----------------------------------------------------------------------------------------------------------------------------------
* Description           : Replacement of strings in xlsx file. 
**********************************************************************************************************************************'''

import openpyxl

#Path
wb = openpyxl.load_workbook(r'sample.xlsx')

#strings to be replaced and replace strings
string1="“You look great today”"
string1r="“You look nice today”"

stringm='''“Hey how are you?

You look different today”'''
stringmr='''“Hello   

How are you doing?”'''

stringc= 'How many different characters do you know? ~,!,@,#,$,%,^,&,*,(,),_,=,+,:,.,”,/,\,|'
stringcr = "A"
#active worksheet data
ws = wb.active

#finding and replacing strings
for i in range(1, ws.max_row + 1):
  for j in range(1, ws.max_column + 1):
      if ws.cell(i,j).value == string1:
         ws.cell(i,j).value = string1r
         print("Found and replaced")
         print(ws.cell(i,j).value)
      if ws.cell(i,j).value == stringm:
         print("Found and replaced")
         ws.cell(i,j).value = stringmr
         print(ws.cell(i,j).value)
         #print(i)
         #print(j)

      #Searching for a string in a string
      if ws.cell(i,j).value == stringc:
         #print(ws.cell(i,j).value.find("How many different characters do you know?"))
         #print("Found wordddddddd")
         #print(type(ws.cell(i,j).value))
         #print(ws.cell(i,j).value.find("~"))
         #i = ws.cell(i,j).value.find("~")
         #ws.cell(i,j).value[i] = "A"

         #Replacing string and storing into the cell
         ws.cell(i,j).value = ws.cell(i,j).value.replace("How many different characters do you know? ~", "A")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("!", "B")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("@", "C")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("#", "D")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("$", "E")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("%", "F")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("^", "G")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("&", "H")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("*", "I")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("(", "J")
         ws.cell(i,j).value = ws.cell(i,j).value.replace(")", "K")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("_", "L")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("=", "M")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("+", "N")
         ws.cell(i,j).value = ws.cell(i,j).value.replace(":", "O")
         ws.cell(i,j).value = ws.cell(i,j).value.replace(".", "P")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("\"", "Q")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("/", "R")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("\\", "S")
         ws.cell(i,j).value = ws.cell(i,j).value.replace("|", "T")

         print("Found and replaced")
         print(ws.cell(i,j).value)
      else:
         print("Not edited")

wb.save("sample.xlsx")
#How many different characters do you know? ~,!,@,#,$,%,^,&,*,(,),_,=,+,:,.,”,/,\,|

#OUTPUT:
# Not edited
# Not edited
# Not edited
# Found and replaced
# “You look nice today”
# Not edited
# Found and replaced
# “Hello   
# 
# How are you doing?”
# Not edited
# Found and replaced
# A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,”,R,S,T
